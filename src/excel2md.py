import os
from markitdown import MarkItDown
from datetime import datetime
import shutil
from openpyxl import load_workbook
from PIL import Image
import io

try:
    import xlrd
    import openpyxl
except ImportError:
    print("请确保已安装 xlrd 和 openpyxl：pip install xlrd openpyxl")
    exit(1)

# 源目录
root_dir = r"d:\Download\audio-visual\heytcm\convert-game-excel-to-md\src\H2-策划案\策划案的文档"
# 统一输出目录
output_dir = r"d:\Download\audio-visual\heytcm\convert-game-excel-to-md\src\md_output"
# 日志文件
log_file = os.path.join(output_dir, "convert.log")
# 临时 xlsx 目录
tmp_xlsx_dir = os.path.join(output_dir, "tmp_xlsx")

excel_exts = ['.xls', '.xlsx']
md = MarkItDown()

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
if not os.path.exists(tmp_xlsx_dir):
    os.makedirs(tmp_xlsx_dir)

def xls_to_xlsx(xls_path, xlsx_path):
    try:
        workbook = xlrd.open_workbook(xls_path)
        new_wb = openpyxl.Workbook()
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            if sheet_idx == 0:
                ws = new_wb.active
                ws.title = sheet.name
            else:
                ws = new_wb.create_sheet(title=sheet.name)
            for row in range(sheet.nrows):
                ws.append(sheet.row_values(row))
        new_wb.save(xlsx_path)
        return True, ''
    except Exception as e:
        return False, str(e)

def extract_images_from_xlsx(xlsx_path, img_output_dir, md_lines, md_base_name, sheet_name=None):
    """
    提取 xlsx 文件中的所有图片到 img_output_dir，并在 md_lines 末尾插入图片引用。
    md_base_name: 当前md文件（不含扩展名），用于图片目录命名
    """
    wb = load_workbook(xlsx_path)
    for sheet in wb.worksheets:
        if sheet_name and sheet.title != sheet_name:
            continue
        if hasattr(sheet, '_images'):
            for idx, image in enumerate(sheet._images):
                img_name = f"{sheet.title}_{idx+1}.png"
                img_path = os.path.join(img_output_dir, img_name)
                try:
                    if hasattr(image, 'image') and isinstance(image.image, Image.Image):
                        image.image.save(img_path)
                    elif hasattr(image, '_data'):
                        img = Image.open(io.BytesIO(image._data()))
                        img.save(img_path)
                    else:
                        continue
                    md_lines.append(f"![](./{os.path.basename(img_output_dir)}/{img_name})\n")
                except Exception as e:
                    md_lines.append(f"<!-- 图片导出失败: {img_name}, 错误: {e} -->\n")

all_excel_files = []

for dirpath, dirnames, filenames in os.walk(root_dir):
    for filename in filenames:
        ext = os.path.splitext(filename)[1].lower()
        file_path = os.path.join(dirpath, filename)
        rel_path = os.path.relpath(file_path, root_dir)
        all_excel_files.append((file_path, rel_path, ext, filename))

# 统计变量
success_count = 0
fail_count = 0
xls2xlsx_success = 0
xls2xlsx_fail = 0
special_char_count = 0

with open(log_file, 'w', encoding='utf-8') as log:
    log.write(f"转换日志 {datetime.now()}\n")
    log.write(f"检测到 {len(all_excel_files)} 个excel文件：\n")
    for file_path, rel_path, ext, filename in all_excel_files:
        log.write(f"  {file_path}\n")
    log.write("\n")
    
    for file_path, rel_path, ext, filename in all_excel_files:
        special_char = any(c in filename for c in '\\/:*?\"<>|')
        if special_char:
            special_char_count += 1
        if ext == '.xls':
            # 转换为 xlsx
            xlsx_rel_path = os.path.splitext(rel_path)[0] + '.xlsx'
            xlsx_full_path = os.path.join(tmp_xlsx_dir, xlsx_rel_path)
            os.makedirs(os.path.dirname(xlsx_full_path), exist_ok=True)
            ok, err = xls_to_xlsx(file_path, xlsx_full_path)
            if ok:
                log.write(f"[xls->xlsx成功] {file_path} -> {xlsx_full_path}\n")
                file_path_for_md = xlsx_full_path
                xls2xlsx_success += 1
            else:
                log.write(f"[xls->xlsx失败] {file_path}, 错误: {err}\n")
                print(f"[xls->xlsx失败] {file_path}, 错误: {err}")
                xls2xlsx_fail += 1
                continue
        else:
            file_path_for_md = file_path
        try:
            result = md.convert(file_path_for_md)
            md_rel_path = os.path.splitext(rel_path)[0] + '.md'
            md_full_path = os.path.join(output_dir, md_rel_path)
            os.makedirs(os.path.dirname(md_full_path), exist_ok=True)
            # 图片目录名为当前md文件名（不含扩展名）
            md_base_name = os.path.splitext(os.path.basename(md_full_path))[0]
            img_output_dir = os.path.join(os.path.dirname(md_full_path), md_base_name)
            os.makedirs(img_output_dir, exist_ok=True)
            md_lines = [result.text_content]
            if file_path_for_md.lower().endswith('.xlsx'):
                extract_images_from_xlsx(file_path_for_md, img_output_dir, md_lines, md_base_name)
            with open(md_full_path, 'w', encoding='utf-8') as f:
                f.writelines(md_lines)
            log.write(f"[成功] {file_path} -> {md_full_path}\n")
            if special_char:
                log.write(f"[警告] 文件名包含特殊字符: {filename}\n")
            print(f"[成功] {file_path} -> {md_full_path}")
            success_count += 1
        except Exception as e:
            log.write(f"[失败] {file_path}, 错误: {e}\n")
            if special_char:
                log.write(f"[警告] 文件名包含特殊字符: {filename}\n")
            print(f"[失败] {file_path}, 错误: {e}")
            fail_count += 1
    # 汇总报告
    log.write("\n========== 转换统计报告 ==========" + "\n")
    log.write(f"总文件数: {len(all_excel_files)}\n")
    log.write(f"成功转换: {success_count}\n")
    log.write(f"失败: {fail_count}\n")
    log.write(f".xls->.xlsx 成功: {xls2xlsx_success}\n")
    log.write(f".xls->.xlsx 失败: {xls2xlsx_fail}\n")
    log.write(f"包含特殊字符的文件: {special_char_count}\n")
    if len(all_excel_files) > 0:
        rate = 100 * success_count / len(all_excel_files)
        log.write(f"整体成功率: {rate:.2f}%\n")
    log.write("==================================\n")
